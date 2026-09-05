"""
Vues Guérite - Entrées/Sorties véhicules
"""
import base64
import json
import os

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count


import os
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

from .models import EnregistrementEntree, BonSortie, EtatBon, StatutEntree, MotifEntree, StatutViewHinstorisue
from .forms import RechercheVehiculeForm, VehiculeForm, ClientForm, ConducteurForm, EntreeForm, SortieForm
from apps.vehicules.models import Marque, Modele, Vehicule, Client, Conducteur
from apps.accounts.decorators import guerite_required 
from apps.audit.service import log_action
from apps.audit.models import ActionType
from itertools import chain
from ..reception.views import dashboard as reception_dashboard
from django.http import Http404, JsonResponse
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from django.http import HttpResponse
from django.db.models import Q

@guerite_required
def dashboard(request):
    # 👉 si réceptionniste → appeler autre view
    if request.user.role == 'RECEPTIONNISTE':
        return reception_dashboard(request)
    
    aujourd_hui = timezone.now().date()
    entrees_today = EnregistrementEntree.objects.filter(date_entree__date=aujourd_hui)
    ctx = { 
        'nb_entrees_today': entrees_today.count(),
        'nb_sorties_today': entrees_today.filter(statut=StatutEntree.SORTI).count(),
        
        'vehicules_presents': EnregistrementEntree.objects.exclude(statut=StatutEntree.SORTI).select_related(
            'vehicule', 'vehicule__client', 'conducteur'
        ).order_by('-date_entree')[:20],
        'entrees_recentes': entrees_today.select_related('vehicule', 'conducteur').order_by('-date_entree')[:10],
    }

    return render(request, 'guerite/dashboard.html', ctx)



@guerite_required
def recherche_vehicule(request):
    form = RechercheVehiculeForm(request.GET or None)
 
    vehicules = []
    query = request.GET.get('q', '').strip()

    if query:
        vehicules = Vehicule.objects.filter(
            Q(immatriculation__icontains=query) |
            Q(marque__icontains=query) |
            Q(modele__icontains=query) |
            Q(client__nom__icontains=query) |
            Q(client__prenom__icontains=query) |
            Q(numero_chassis__icontains=query)
        ).select_related('client').order_by('immatriculation')[:20]

# .exclude(
            # 👉 EXCLURE ceux qui sont EN_COURS
           #  entrees__statut=StatutEntree.EN_COURS  )

    return render(request, 'guerite/recherche_vehicule.html', {
        'form': form,
        'vehicules': vehicules,
        'query': query
    })


@guerite_required
def nouvelle_entree(request):
    """Étape 1: choisir un véhicule existant ou en créer un."""
    vehicule_id = request.GET.get('vehicule_id')
    vehicule = None
    if vehicule_id:
        vehicule = get_object_or_404(Vehicule, id=vehicule_id)
    return render(request, 'guerite/entree/choix.html', {'vehicule': vehicule})



# ─── AJAX: Autocomplete conducteurs ──────────────────────────────────────────

@login_required
def autocomplete_conducteurs(request):
    q = request.GET.get('q', '').strip()
    results = []

    # 1. Recherche prioritaire dans la table Conducteur
    qs_conducteurs = Conducteur.objects.all()
    if q:
        qs_conducteurs = qs_conducteurs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) |
            Q(telephone__icontains=q) | Q(cni__icontains=q) |
            Q(permis__icontains=q)
        )

    for c in qs_conducteurs[:10]:
        nom = f"{c.prenom} {c.nom}".strip()
        detail_parts = []
        if c.telephone:
            detail_parts.append(c.telephone)
        if c.permis:
            detail_parts.append(f"Permis: {c.permis}")
        if c.categorie_permis:
            detail_parts.append(f"Cat. {c.categorie_permis}")

        results.append({
            'id': str(c.id),
            'nom': nom,
            'detail': ' — '.join(detail_parts),
            'is_client': False,
        })

    # 2. Si moins de 10 résultats, recherche complémentaire dans les Clients Particuliers
    limit_restante = 10 - len(results)
    if limit_restante > 0:
        # Ajustez 'type_client' ou 'type' selon le champ exact de votre modèle Client
        qs_clients = Client.objects.filter(is_active=True, type_client='PARTICULIER')

        if q:
            qs_clients = qs_clients.filter(
                Q(nom__icontains=q) | Q(prenom__icontains=q) |
                Q(telephone__icontains=q) 
            )

        for client in qs_clients[:limit_restante]:
            nom = f"{getattr(client, 'prenom', '')} {getattr(client, 'nom', '')}".strip()
            detail_parts = ["Client Particulier (À créer)"]
            
            if getattr(client, 'telephone', None):
                detail_parts.append(client.telephone)

            results.append({
                'id': f"client_{client.id}",  # Identifiant distinct pour le frontend
                'client_id': client.id,
                'nom': nom,
                'prenom': getattr(client, 'prenom', ''),
                'nom_famille': getattr(client, 'nom', ''),
                'telephone': getattr(client, 'telephone', ''),
                'cni': getattr(client, 'cni', ''),
                'detail': ' — '.join(detail_parts),
                'is_client': True,  # Flag indiquant au JS qu'il faut créer un conducteur
            })

    return JsonResponse({'results': results})


def autocomplete_clients(request):
    """Retourne la liste des clients filtrés pour l'autocomplete."""

    q    = request.GET.get('q', '').strip()
    type_client = request.GET.get('type', '')
    qs = Client.objects.filter(is_active=True)
    if type_client in ('PARTICULIER', 'ENTREPRISE'):
        qs = qs.filter(type_client=type_client)
    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) 

        )
    results = []
    for c in qs[:10]:
        if c.type_client == 'PARTICULIER':
            detail = f"{c.telephone} — {c.ville or 'Particulier'}"
            nom = f"{c.prenom} {c.nom}".strip() if c.prenom else c.nom
        else:
            detail = f"{c.telephone} — Entreprise{' | ' + c.ninea if c.ninea else ''}"
            nom = c.nom
        results.append({'id': str(c.id), 'nom': nom, 'detail': detail})
    from django.http import JsonResponse
    return JsonResponse({'results': results})




@guerite_required 
def creer_client(request):
    form = ClientForm()
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            client.created_by = request.user
            client.save()
            log_action(request, ActionType.CREATION, 'GUERITE', client)
            messages.success(request, f"Client {client} créé.")

        
    return render(request, 'guerite/entree/creer_client.html', {'form': form})



def creer_client_ajax(request):
    """Créer un client via appel AJAX depuis le formulaire véhicule."""
 
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
 
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Données invalides'})
 
    type_client = data.get('type_client', 'PARTICULIER')
    nom         = data.get('nom', '').strip()
    telephone   = data.get('telephone', '').strip()
 
    if not nom or not telephone:
        return JsonResponse({'success': False, 'error': 'Nom et téléphone obligatoires'})
 
    if Client.objects.filter(telephone=telephone).exists():
        existing = Client.objects.get(telephone=telephone)
        return JsonResponse({
            'success': True,
            'id': str(existing.id),
            'nom': str(existing),
            'detail': f"{existing.telephone} — {existing.ville or ''}",
        })
 
    client = Client.objects.create(
        type_client = type_client,
        nom         = nom,
        prenom      = data.get('prenom', ''),
        nom_correspondant= data.get('nom_correspondant', ''),
        telephone   = telephone,
        telephone2  = data.get('telephone2', ''),
        email       = data.get('email', ''),
        adresse     = data.get('adresse', ''),
        ville       = data.get('ville', ''),
        ninea       = data.get('ninea', ''),
        created_by  = request.user,
    )
    log_action(request, ActionType.CREATION, 'GUERITE', client)
    return JsonResponse({
        'success': True,
        'id':     str(client.id),
        'nom':    str(client),
        'detail': f"{client.telephone} — {client.ville or type_client}",
    })




def autocomplete_marques(request):
    term = request.GET.get('q', '').strip()
    if len(term) < 2:
        return JsonResponse({'results': []})
    marques = Marque.objects.filter(nom__icontains=term)[:15]
    results = [{'id': m.id, 'nom': m.nom} for m in marques]
    return JsonResponse({'results': results})



def autocomplete_modeles(request):
    q = request.GET.get('q', '')
    marque = request.GET.get('marque', '')
    modeles = Modele.objects.filter(nom__icontains=q)
    if marque:
        modeles = modeles.filter(marque__nom__icontains=marque)
    modeles = modeles[:10]
    results = [{'nom': m.nom} for m in modeles]
    return JsonResponse({'results': results})


def creer_conducteur_ajax(request):
    """Créer un conducteur via appel AJAX depuis le formulaire véhicule."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Données invalides'})

    # Champs obligatoires
    nom = data.get('nom', '').strip()
    telephone = data.get('telephone', '').strip()
    if not nom or not telephone:
        return JsonResponse({'success': False, 'error': 'Nom et téléphone obligatoires'})

    # Vérifier si un conducteur avec ce téléphone existe déjà
    conducteur_existant = Conducteur.objects.filter(telephone=telephone).first()
    if conducteur_existant:
        return JsonResponse({
            'success': True,
            'id': str(conducteur_existant.id),
            'nom': str(conducteur_existant),
            'detail': f"{conducteur_existant.telephone} — {conducteur_existant.ville or ''}",
        })

    # Création du nouveau conducteur
    conducteur = Conducteur.objects.create(
        nom=nom,
        prenom=data.get('prenom', '').strip(),
        telephone=telephone,
        telephone2=data.get('telephone2', '').strip(),
        cni=data.get('cni', '').strip(),
        permis=data.get('permis', '').strip(),
        categorie_permis=data.get('categorie_permis', '').strip(),
    )
    log_action(request, ActionType.CREATION, 'GUERITE', conducteur)
    return JsonResponse({
        'success': True,
        'id': str(conducteur.id),
        'nom': str(conducteur),
        'detail': f"{conducteur.telephone}",
    })



@guerite_required
def creer_conducteur(request):
    form = ConducteurForm()
    if request.method == 'POST':
        form = ConducteurForm(request.POST)
        if form.is_valid():
            conducteur = form.save(commit=False)
            conducteur.created_by = request.user
            conducteur.save()
            log_action(request, ActionType.CREATION, 'GUERITE', conducteur)
            messages.success(request, f"Conducteur {conducteur} créé.")
            return redirect(request.GET.get('next', 'nouvelle_entree'))
    return render(request, 'guerite/entree/creer_conducteur.html', {'form': form})


@guerite_required 
def creer_vehicule(request):
    """Version améliorée avec type propriétaire et photos multiples."""
    client_id = request.GET.get('client_id')
    client_preselect = None
    if client_id:
        try:
            c = Client.objects.get(id=client_id)
            client_preselect = str(c)
        except Client.DoesNotExist:
            client_id = None
 
    form = VehiculeForm(initial={'client': client_id} if client_id else {})
 
    if request.method == 'POST':
        form = VehiculeForm(request.POST, request.FILES)
        # Validation manuelle du client (sélectionné via AJAX)
        client_id_post = request.POST.get('client')
        if not client_id_post:
            form.add_error('client', 'Veuillez sélectionner un propriétaire.')
 
        if form.is_valid():
            vehicule = form.save(commit=False)
            vehicule.created_by = request.user

            # Récupération ou création de la marque
            marque_nom = form.cleaned_data['marque'].strip()
            marque, _ = Marque.objects.get_or_create(nom=marque_nom)
            
            # Récupération ou création du modèle (lié à cette marque)
            modele_nom = form.cleaned_data['modele'].strip()

            modele, _ = Modele.objects.get_or_create(
                nom=modele_nom,
                marque=marque
            )

            # Immatriculation en majuscules 
            vehicule.immatriculation = vehicule.immatriculation.upper()
            vehicule.save()
 
        
            photos_paths = []
            for i in range(3):
                 photo_file = request.FILES.get(f'photo_{i}')
                 if photo_file:
                      ext = os.path.splitext(photo_file.name)[1]
                      immat_clean = vehicule.immatriculation.replace(' ', '').replace('-', '').upper()
                      filename = f"vehicules/photos/{immat_clean}_{i}{ext}"
                      saved_path = default_storage.save(filename, ContentFile(photo_file.read()))
                      photos_paths.append(saved_path)

                      if photos_paths:
                         vehicule.photos = ';'.join(photos_paths)
                         vehicule.photo = photos_paths[0]
                         vehicule.save(update_fields=['photos', 'photo'])
               
            log_action(request, ActionType.CREATION, 'GUERITE', vehicule)
            messages.success(request, f"Véhicule {vehicule.immatriculation} créé.")
            return redirect(f"/guerite/entree/enregistrer/?vehicule_id={vehicule.id}")
        
    return render(request, 'guerite/entree/creer_vehicule.html', { 
        'form':             form,
        'client_id':        client_id,
        'client_preselect': client_preselect,
    })


@guerite_required
def enregistrer_entree(request):
    vehicule_id = request.GET.get('vehicule_id')
    vehicule = None
    if vehicule_id:
        vehicule = get_object_or_404(Vehicule, id=vehicule_id)

    initial = {'vehicule': vehicule} if vehicule else {}
    form = EntreeForm(initial=initial)

    if request.method == 'POST':
        form = EntreeForm(request.POST)
        if form.is_valid():
            entree = form.save(commit=False)
            entree.agent_entree = request.user
            entree.save()
            log_action(request, ActionType.CREATION, 'GUERITE', entree, {'motif': entree.motif})
            messages.success(request, f"Entrée {entree.numero} enregistrée avec succès !")
            return redirect('detail_entree', entree_id=entree.id)

    conducteurs = sorted(
    chain(
        Conducteur.objects.all(),
        Client.objects.all()
    ),
    key=lambda x: x.nom
)
    return render(request, 'guerite/entree/enregistrer.html', {
        'form': form, 'vehicule': vehicule, 'conducteurs': conducteurs
    })


@guerite_required
def detail_entree(request, entree_id):
    entree = get_object_or_404(
        EnregistrementEntree.objects.select_related('vehicule', 'vehicule__client', 'conducteur', 'agent_entree'),
        id=entree_id
    )
    return render(request, 'guerite/entree/detail.html', {'entree': entree})
 

@guerite_required
def modifier_motif_entree(request, entree_id):
    entree = get_object_or_404(EnregistrementEntree, pk=entree_id)

    if request.method == "POST":
        motif = request.POST.get("motif")

        if motif in ["REPARATION", "VISITE"]:
            entree.motif = motif
            entree.save()

            messages.success(
                request,
                f"Motif modifié : {entree.get_motif_display()}"
            )

    return redirect("detail_entree", entree_id=entree.id)


@guerite_required
def liste_vehicules_presents(request):
    entrees = (
        EnregistrementEntree.objects.exclude(statut=StatutEntree.SORTI)
        .select_related(
            'vehicule',
            'vehicule__client',
            'conducteur',
            'conducteur__client',  # Inclus le client rattaché au conducteur
        )
        .order_by('-date_entree')
    )
    return render(request, 'guerite/vehicules_presents.html', {'entrees': entrees})


def get_entree_non_sortie(entree_id):
    entree = get_object_or_404(EnregistrementEntree, id=entree_id)
    if entree.statut == StatutEntree.SORTI:
        raise Http404("Entrée déjà sortie")
    return entree

# Utilisation


@guerite_required
def enregistrer_sortie(request, entree_id):
    entree = get_entree_non_sortie(entree_id)
    bon_sortie = None

    # Recherche du bon de sortie uniquement pour une réparation
    if entree.motif == MotifEntree.REPARATION:
        bons = BonSortie.objects.filter(
            vehicule=entree.vehicule,
            est_valide=False
        )
        if bons.exists():
            bon_sortie = bons.latest('created_at')

    form = SortieForm()

    if request.method == 'POST':
        form = SortieForm(request.POST)

        # Vérification du bon de sortie pour les réparations
        if entree.motif == MotifEntree.REPARATION and not bon_sortie:
            messages.error(
                request,
                "Aucun bon de sortie valide trouvé. Contactez la réception."
            )
            return render(
                request,
                'guerite/sortie/enregistrer.html',
                {
                    'entree': entree,
                    'form': form,
                    'bon_sortie': bon_sortie
                }
            )

        # Enregistrement de la sortie
        entree.statut = StatutEntree.SORTI
        entree.date_sortie = timezone.now()
        entree.agent_sortie = request.user

        # Mise à jour du bon uniquement pour une réparation
        if entree.motif == MotifEntree.REPARATION and bon_sortie:
            bon_sortie.est_valide = True
            bon_sortie.valide_par = request.user
            bon_sortie.date_validation = timezone.now()
            bon_sortie.etats = EtatBon.VALIDER
            bon_sortie.save()

            entree.bon_sortie = bon_sortie

        entree.save()

        log_action(
            request,
            ActionType.CHANGEMENT_STATUT,
            'GUERITE',
            entree,
            {'statut': 'SORTI'}
        )

        messages.success(
            request,
            f"Sortie du véhicule {entree.vehicule.immatriculation} enregistrée."
        )

        return redirect('dashboard_guerite')

    return render(
        request,
        'guerite/sortie/enregistrer.html',
        {
            'entree': entree,
            'form': form,
            'bon_sortie': bon_sortie
        }
    )

@guerite_required
def consulter_bon_sortie(request):
    numero = request.GET.get('numero', '').strip()
    bon = None
    if numero:
        try:
            bon = BonSortie.objects.select_related('vehicule', 'vehicule__client').get(numero=numero)
        except BonSortie.DoesNotExist:
            messages.warning(request, f"Aucun bon de sortie trouvé avec le numéro {numero}")
    return render(request, 'guerite/bon_sortie/consulter.html', {'bon': bon, 'numero': numero})


@guerite_required 
def historique_entrees(request):
    entrees = get_filtered_entrees(request)
    return render(request, 'guerite/historique.html', {
        'entrees': entrees[:100],
        'statuts': StatutViewHinstorisue.choices,
        'motifs': MotifEntree.choices,
    })

def get_filtered_entrees(request):
    entrees = EnregistrementEntree.objects.select_related(
        'vehicule', 'vehicule__client', 'conducteur', 'agent_entree'
    ).order_by('-date_entree')

    statut = request.GET.get('statut')
    motif = request.GET.get('motif')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    q = request.GET.get('q', '')


    if statut == 'SORTI':
        entrees = entrees.filter(statut='SORTI')
    if statut == 'PRESENT':
        entrees = entrees.exclude(statut='SORTI')    
    if motif:
        entrees = entrees.filter(motif=motif)
    if date_debut:
        entrees = entrees.filter(date_entree__date__gte=date_debut)
    if date_fin:
        entrees = entrees.filter(date_entree__date__lte=date_fin)
    if q:
        entrees = entrees.filter(
            Q(vehicule__immatriculation__icontains=q) |
            Q(numero__icontains=q) |
            Q(vehicule__client__nom__icontains=q)
        )
    return entrees  # retourne tout, pas de limite 100




def export_entrees_excel(request):
    # Récupérer les mêmes filtres que dans historique_entrees
    entrees = get_filtered_entrees(request)  # on factorise la logique

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique entrées"

    # En-têtes
    headers = ['N°', 'Immatriculation', 'Client', 'Téléphone client', 'Conducteur', 'Motif', 'Statut', 'Date entrée', 'Date sortie', 'Agent entrée', 'Observations']
    ws.append(headers)
    # Mise en forme des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    for e in entrees:
        ws.append([
            e.numero,
            e.vehicule.immatriculation,
            str(e.vehicule.client),
            e.vehicule.client.telephone if e.vehicule.client else '',
            str(e.conducteur) if e.conducteur else '',
            e.get_motif_display(),
            e.get_statut_display(),
            e.date_entree.strftime("%d/%m/%Y %H:%M") if e.date_entree else '',
            e.date_sortie.strftime("%d/%m/%Y %H:%M") if e.date_sortie else '',
            e.agent_entree.full_name if e.agent_entree else '',
            e.observations or ''
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historique_entrees_{datetime.date.today()}.xlsx'
    wb.save(response)
    return response


import csv

def export_entrees_csv(request):
    entrees = get_filtered_entrees(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=historique_entrees_{datetime.date.today()}.csv'
    writer = csv.writer(response)
    writer.writerow(['N°', 'Immatriculation', 'Client', 'Conducteur', 'Motif', 'Statut', 'Date entrée', 'Date sortie', 'Agent'])
    for e in entrees:
        writer.writerow([
            e.numero, e.vehicule.immatriculation, str(e.vehicule.client), str(e.conducteur),
            e.get_motif_display(), e.get_statut_display(),
            e.date_entree.strftime("%d/%m/%Y %H:%M") if e.date_entree else '',
            e.date_sortie.strftime("%d/%m/%Y %H:%M") if e.date_sortie else '',
            e.agent_entree.full_name if e.agent_entree else ''
        ])
    return response


def export_entrees_pdf(request):
    entrees = get_filtered_entrees(request)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=historique_entrees_{datetime.date.today()}.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    elements = []

    # Titre
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1  # centre
    elements.append(Paragraph("Historique des entrées", title_style))
    elements.append(Spacer(1, 12))

    # Sous-titre (filtres appliqués)
    filters = []
    if request.GET.get('q'):
        filters.append(f"Recherche: {request.GET['q']}")
    if request.GET.get('statut'):
        filters.append(f"Statut: {dict(StatutEntree.choices).get(request.GET['statut'], '')}")
    if request.GET.get('motif'):
        filters.append(f"Motif: {dict(MotifEntree.choices).get(request.GET['motif'], '')}")
    if request.GET.get('date_debut'):
        filters.append(f"Du: {request.GET['date_debut']}")
    if request.GET.get('date_fin'):
        filters.append(f"Au: {request.GET['date_fin']}")
    if filters:
        filter_text = "Filtres : " + ", ".join(filters)
        filter_style = ParagraphStyle('FilterStyle', parent=styles['Normal'], fontSize=9, textColor=colors.gray)
        elements.append(Paragraph(filter_text, filter_style))
        elements.append(Spacer(1, 10))

    # Données du tableau
    data = [['Immat.', 'Client', 'Motif', 'Statut', 'Entrée', 'Sortie', 'Agent']]
    for e in entrees:
        data.append([
            e.vehicule.immatriculation,
            str(e.vehicule.client)[:30],
            e.get_motif_display(),
            e.get_statut_display(),
            e.date_entree.strftime("%d/%m %H:%M") if e.date_entree else '',
            e.date_sortie.strftime("%d/%m %H:%M") if e.date_sortie else '',
            e.agent_entree.full_name.split()[0] if e.agent_entree else ''
        ])

    # Création du tableau
    table = Table(data, repeatRows=1, colWidths=[ 70, 100, 60, 55, 70, 70, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    return response